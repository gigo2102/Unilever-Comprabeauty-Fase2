"""
Script para generar reporte detallado de grupos de geo-localización
"""
import pandas as pd
import os

# Obtener rutas
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_DIR = os.path.dirname(SCRIPT_DIR)
CSV_PATH = os.path.join(PROJECT_DIR, "Csvs")
DATOS_PATH = os.path.join(PROJECT_DIR, "datos_generados")

print("=" * 80)
print("GENERANDO REPORTE DETALLADO DE GRUPOS GEO-LOCALIZACIÓN")
print("=" * 80)
print()

# Cargar datos
print("Cargando datos...")
locations = pd.read_csv(os.path.join(CSV_PATH, "CustomerLocation.csv"))
users = pd.read_csv(os.path.join(CSV_PATH, "EcommerceUsers.csv"))
fase2 = pd.read_csv(
    os.path.join(DATOS_PATH, "fase2_consolidacion_geo_proximidad.csv"),
    encoding='utf-8-sig'
)

print(f"Grupos en Fase 2: {fase2['grupo_consolidacion_id'].nunique()}")
print()

# Merge con locations para obtener detalles
print("Combinando con datos de locations...")
fase2_detalle = fase2.merge(
    locations[['ID', 'FantasyName', 'TaxId', 'FullAddress', 'Latitude', 'Longitude',
               'EcommerceUserCreatorId', 'CreatedDate', 'Group', 'Region', 'Zone']],
    left_on='location_id',
    right_on='ID',
    how='left'
)

# Merge con usuarios para obtener email del creador
print("Combinando con datos de usuarios...")
fase2_detalle = fase2_detalle.merge(
    users[['ID', 'Email', 'FullName']],
    left_on='EcommerceUserCreatorId',
    right_on='ID',
    how='left',
    suffixes=('', '_user')
)

# Ordenar por grupo y master primero
fase2_detalle = fase2_detalle.sort_values(
    ['grupo_consolidacion_id', 'is_master'],
    ascending=[True, False]
)

# Seleccionar columnas relevantes
columnas_reporte = [
    'grupo_consolidacion_id',
    'location_id',
    'is_master',
    'FantasyName',
    'TaxId',
    'FullAddress',
    'Latitude',
    'Longitude',
    'Email',
    'FullName',
    'Group',
    'Region',
    'Zone',
    'CreatedDate',
    'confianza',
    'alerta',
    'locations_en_grupo',
    'taxids_unicos',
    'distancia_max_metros'
]

reporte = fase2_detalle[columnas_reporte].copy()

# Renombrar columnas para mejor legibilidad
reporte.columns = [
    'Grupo ID',
    'Location ID',
    'Es Master',
    'Nombre Location',
    'TaxID',
    'Dirección',
    'Latitud',
    'Longitud',
    'Email Creador',
    'Nombre Creador',
    'Grupo',
    'Región',
    'Zona',
    'Fecha Creación',
    'Confianza',
    'Alerta',
    'Locations en Grupo',
    'TaxIDs Únicos',
    'Distancia Máxima (m)'
]

# Guardar a Excel con formato
output_file = os.path.join(DATOS_PATH, "reporte_grupos_geolocalizacion_detallado.xlsx")
print(f"Guardando reporte en {output_file}...")

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment

with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    reporte.to_excel(writer, sheet_name='Grupos Geo', index=False)

    # Aplicar formato
    workbook = writer.book
    worksheet = writer.sheets['Grupos Geo']

    # Formato de encabezado
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Formato para filas master
    master_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
    master_font = Font(bold=True)

    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
        # Si es master (columna C)
        if row[2].value == True:
            for cell in row:
                cell.fill = master_fill
                cell.font = master_font

    # Ajustar ancho de columnas
    worksheet.column_dimensions['A'].width = 15  # Grupo ID
    worksheet.column_dimensions['B'].width = 12  # Location ID
    worksheet.column_dimensions['C'].width = 10  # Es Master
    worksheet.column_dimensions['D'].width = 30  # Nombre
    worksheet.column_dimensions['E'].width = 20  # TaxID
    worksheet.column_dimensions['F'].width = 40  # Dirección
    worksheet.column_dimensions['G'].width = 12  # Latitud
    worksheet.column_dimensions['H'].width = 12  # Longitud
    worksheet.column_dimensions['I'].width = 35  # Email
    worksheet.column_dimensions['J'].width = 25  # Nombre Creador
    worksheet.column_dimensions['K'].width = 15  # Grupo
    worksheet.column_dimensions['L'].width = 15  # Región
    worksheet.column_dimensions['M'].width = 15  # Zona
    worksheet.column_dimensions['N'].width = 20  # Fecha
    worksheet.column_dimensions['O'].width = 12  # Confianza
    worksheet.column_dimensions['P'].width = 25  # Alerta
    worksheet.column_dimensions['Q'].width = 12  # Locations
    worksheet.column_dimensions['R'].width = 12  # TaxIDs Únicos
    worksheet.column_dimensions['S'].width = 18  # Distancia

    # Congelar primera fila
    worksheet.freeze_panes = 'A2'

print()
print("=" * 80)
print("REPORTE GENERADO EXITOSAMENTE")
print("=" * 80)
print(f"Total de registros: {len(reporte)}")
print(f"Total de grupos: {reporte['Grupo ID'].nunique()}")
print(f"Archivo: {output_file}")
print()
